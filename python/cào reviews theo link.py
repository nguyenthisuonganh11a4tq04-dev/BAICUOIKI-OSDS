import os
import time
from datetime import datetime
from typing import Optional, List, Tuple

import dateparser
from pymongo import MongoClient, UpdateOne
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# ===================== CONFIG =====================
MONGO_URI = "mongodb://localhost:27017/"
DB_NAME = "reviews-db"
EXCEL_PATH = "restaurants_all_districts_from_home_1.xlsx"  # path tới file Excel
EXCEL_LINK_HEADER_CANDIDATES = ["link", "Link", "URL", "url", "Đường dẫn", "Link quán", "Restaurant URL"]

MAX_REVIEW_PAGES = 20  # giới hạn số trang review duyệt
HEADLESS = True
LANG = "vi-VN"

FOODY_BASE = "https://www.foody.vn"
LOGIN_URL = "https://www.foody.vn/account/login"

# Điền thông tin đăng nhập (hoặc dùng biến môi trường)
FOODY_EMAIL = os.getenv("FOODY_EMAIL", "your_email@example.com")
FOODY_PASSWORD = os.getenv("FOODY_PASSWORD", "your_password")

# ===================== DB =====================
client = MongoClient(MONGO_URI)
db = client[DB_NAME]
restaurants_col = db["restaurants"]
foods_col = db["foods"]
reviews_col = db["reviews"]

# ===================== UTILS =====================
def now_date_str():
    return datetime.now().strftime("%Y-%m-%d")

def to_iso(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    dt = dateparser.parse(
        text,
        languages=["vi", "en"],
        settings={"TIMEZONE": "Asia/Ho_Chi_Minh", "RETURN_AS_TIMEZONE_AWARE": False}
    )
    return dt.isoformat() if dt else None

def setup_driver(headless=True) -> webdriver.Chrome:
    opts = Options()
    if headless:
        # new headless mode (Chrome >= 109)
        opts.add_argument("--headless=new")
    opts.add_argument("--window-size=1366,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument(f"--lang={LANG}")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome Safari")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    driver.set_page_load_timeout(60)
    return driver

def safe_text(el) -> str:
    try:
        return el.text.strip()
    except:
        return ""

def find_or_none(parent, by, value):
    try:
        return parent.find_element(by, value)
    except NoSuchElementException:
        return None

def wait_css(driver, selector, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.CSS_SELECTOR, selector)))

def wait_click(driver, selector, timeout=10):
    el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.CSS_SELECTOR, selector)))
    driver.execute_script("arguments[0].click();", el)
    return el

# ===================== LOGIN =====================
def login_foody(driver):
    driver.get(LOGIN_URL)
    try:
        # Chờ form đăng nhập xuất hiện
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, "form#form_login, form[action*='login']")))
    except TimeoutException:
        return

    # Thử nhiều selector phổ biến cho email/password
    email_selectors = ["input#email", "input[name='email']", "input[type='email']"]
    pass_selectors = ["input#password", "input[name='password']", "input[type='password']"]
    submit_selectors = ["button[type='submit']", "button.btn-login", "input[type='submit']"]

    email_el = None
    for sel in email_selectors:
        email_el = find_or_none(driver, By.CSS_SELECTOR, sel)
        if email_el:
            break

    pass_el = None
    for sel in pass_selectors:
        pass_el = find_or_none(driver, By.CSS_SELECTOR, sel)
        if pass_el:
            break

    if not email_el or not pass_el:
        return

    # Điền thông tin đăng nhập
    email_el.clear()
    email_el.send_keys(FOODY_EMAIL)
    pass_el.clear()
    pass_el.send_keys(FOODY_PASSWORD)

    # Click nút đăng nhập
    clicked = False
    for sel in submit_selectors:
        btn = find_or_none(driver, By.CSS_SELECTOR, sel)
        if btn:
            driver.execute_script("arguments[0].click();", btn)
            clicked = True
            break

    if not clicked:
        # Nhấn Enter nếu không tìm thấy nút
        pass_el.submit()

    # Chờ chuyển trang/đăng nhập xong (nếu có captcha thì cần thao tác tay)
    time.sleep(3)

# ===================== EXCEL LINKS =====================
def read_links_from_excel(path: str) -> List[str]:
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    # Tìm cột link theo header candidates
    link_col_idx = None
    for i, h in enumerate(headers):
        h_str = str(h).strip()
        if h_str in EXCEL_LINK_HEADER_CANDIDATES:
            link_col_idx = i
            break

    # Nếu không có header phù hợp, lấy cột đầu tiên
    if link_col_idx is None:
        link_col_idx = 0

    links = []
    for row in ws.iter_rows(min_row=2):
        cell = row[link_col_idx]
        val = cell.value
        if not val:
            continue
        url = str(val).strip()
        if url.startswith("http"):
            links.append(url)
    wb.close()
    return links

# ===================== RESTAURANT + FOODS =====================
def crawl_restaurant_and_foods(driver, url) -> Tuple[dict, List[dict]]:
    driver.get(url)
    # Đảm bảo trang đã tải tiêu đề
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "h1")))
    except TimeoutException:
        pass
    time.sleep(1.0)

    name = safe_text(find_or_none(driver, By.CSS_SELECTOR, "h1"))
    address = safe_text(find_or_none(driver, By.CSS_SELECTOR, ".res-common-add, .rd-address, .rd-addr"))

    res_doc = {
        "url": url,
        "name": name,
        "address": address,
        "first_seen": now_date_str(),
        "last_updated": now_date_str()
    }

    foods = []
    dish_els = driver.find_elements(By.CSS_SELECTOR, ".menu-item-name, .dish-name, .txt-menu-item")
    for d in dish_els:
        food_name = safe_text(d)
        if 3 < len(food_name) < 200:
            foods.append({
                "restaurant_url": url,
                "food_name": food_name,
                "crawl_date": now_date_str()
            })

    return res_doc, foods

# ===================== REVIEWS =====================
def get_latest_review_time(url) -> Optional[str]:
    r = reviews_col.find_one({"restaurant_url": url}, sort=[("comment_time", -1)])
    return r["comment_time"] if r else None

def goto_review_tab(driver):
    # Thử bấm vào tab chứa review/bình luận
    labels = ["Đánh giá", "Bình luận", "Reviews"]
    for label in labels:
        try:
            tab = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, label)))
            driver.execute_script("arguments[0].click();", tab)
            time.sleep(1.2)
            return True
        except:
            continue
    return False

def save_review_immediately(doc: dict):
    # Lưu từng review ngay (tới đâu lưu tới đó) bằng upsert
    reviews_col.update_one(
        {
            "restaurant_url": doc["restaurant_url"],
            "comment_text": doc["comment_text"],
            "comment_time": doc["comment_time"]
        },
        {"$set": doc},
        upsert=True
    )

def crawl_reviews_incremental(driver, url, max_pages=MAX_REVIEW_PAGES):
    latest_time = get_latest_review_time(url)

    # Vào trang quán
    driver.get(url)
    time.sleep(1.2)
    goto_review_tab(driver)

    # Duyệt từng trang review, giả định trang đầu là mới nhất (giảm dần theo thời gian)
    for page in range(max_pages):
        # Chờ các item xuất hiện
        try:
            WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, ".review-item, .rd-item"))
            )
        except TimeoutException:
            break

        items = driver.find_elements(By.CSS_SELECTOR, ".review-item, .rd-item")
        if not items:
            break

        stop = False
        for it in items:
            # Nội dung
            content = safe_text(find_or_none(it, By.CSS_SELECTOR, ".review-text, .rd-des, .rdes"))
            # Thời gian
            time_str = safe_text(find_or_none(it, By.CSS_SELECTOR, ".time, .date, .rd-time"))
            iso_time = to_iso(time_str)

            # Nếu đã có mốc mới nhất, gặp review cũ hơn hoặc bằng thì dừng
            if latest_time and iso_time and iso_time <= latest_time:
                stop = True
                break

            if content:
                doc = {
                    "restaurant_url": url,
                    "comment_text": content,
                    "comment_time": iso_time,
                    "crawl_date": now_date_str()
                }
                # Lưu ngay vào MongoDB
                save_review_immediately(doc)

        if stop:
            break

        # Sang trang tiếp theo
        next_btn = find_or_none(driver, By.CSS_SELECTOR, "a.next, a[rel='next'], .pagination .next")
        if not next_btn:
            break
        driver.execute_script("arguments[0].click();", next_btn)
        time.sleep(1.2)

# ===================== SAVE RESTAURANT + FOODS =====================
def save_restaurant_and_foods(res: dict, foods: List[dict]):
    restaurants_col.update_one({"url": res["url"]}, {"$set": res}, upsert=True)

    if foods:
        foods_col.bulk_write([
            UpdateOne(
                {"restaurant_url": f["restaurant_url"], "food_name": f["food_name"]},
                {"$set": f},
                upsert=True
            ) for f in foods
        ], ordered=False)

# ===================== MAIN =====================
def run():
    driver = setup_driver(headless=HEADLESS)
    try:
        # Đăng nhập trước (nếu cần xem đầy đủ review)
        login_foody(driver)

        # Đọc links từ file Excel (streaming, không cần nạp toàn bộ vào RAM)
        links = read_links_from_excel(EXCEL_PATH)
        print(f"Đọc được {len(links)} link từ Excel")

        for i, url in enumerate(links, start=1):
            print(f"[{i}/{len(links)}] {url}")
            try:
                # Cào thông tin quán và menu
                res, foods = crawl_restaurant_and_foods(driver, url)
                save_restaurant_and_foods(res, foods)

                # Cào review theo thời gian giảm dần (incremental, tới đâu lưu tới đó)
                crawl_reviews_incremental(driver, url, MAX_REVIEW_PAGES)

            except Exception as e:
                print(f"   Lỗi: {e}")
                continue

    finally:
        driver.quit()

if __name__ == "__main__":
    run()